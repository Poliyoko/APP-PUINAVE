"""SPT-001B-P07 v1.1: normalización auditable de encabezados."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .excel_reader import LectorExcelRLB
from .exporter import exportar_resultado
from .schema_loader import cargar_esquema


@dataclass(frozen=True, slots=True)
class SugerenciaEncabezado:
    original: str
    normalizado: str
    campo_canonico: str | None
    alias_recomendado: str | None
    confianza: float
    decision: str


def normalizar_texto(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    without_accents = "".join(
        char
        for char in decomposed
        if not unicodedata.combining(char)
    )
    lowered = without_accents.casefold()
    cleaned = re.sub(r"[^a-z0-9]+", " ", lowered)
    return " ".join(cleaned.split())


def _tokens(value: str) -> set[str]:
    return set(normalizar_texto(value).split())


def similitud(a: str, b: str) -> float:
    """Calcula similitud evitando coincidencias dentro de otras palabras."""

    na = normalizar_texto(a)
    nb = normalizar_texto(b)

    if not na or not nb:
        return 0.0

    if na == nb:
        return 1.0

    ta = _tokens(na)
    tb = _tokens(nb)

    sequence = SequenceMatcher(None, na, nb).ratio()

    token_score = 0.0
    if ta or tb:
        token_score = len(ta & tb) / len(ta | tb)

    exact_token_containment = bool(ta) and (
        ta.issubset(tb) or tb.issubset(ta)
    )

    containment_score = 0.92 if exact_token_containment else 0.0

    combined = (sequence * 0.65) + (token_score * 0.35)

    return round(
        max(
            sequence,
            combined,
            containment_score,
        ),
        4,
    )


def leer_encabezados_excel(
    excel: str | Path,
    max_rows: int = 20,
) -> dict[str, Any]:
    path = Path(excel)

    if not path.is_file():
        raise FileNotFoundError(
            f"No se encontró el Excel: {path}"
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        result: dict[str, Any] = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            best_row = None
            best_values: list[str] = []
            best_score = -1

            limit = min(sheet.max_row, max_rows)

            for row_number, values in enumerate(
                sheet.iter_rows(
                    min_row=1,
                    max_row=limit,
                    values_only=True,
                ),
                start=1,
            ):
                headers = [
                    str(value).strip()
                    for value in values
                    if value is not None
                    and str(value).strip()
                ]

                if not headers:
                    continue

                score = len(headers)
                terms = " ".join(
                    normalizar_texto(value)
                    for value in headers
                )

                for keyword in (
                    "puinave",
                    "espanol",
                    "ingles",
                    "palabra",
                    "traduccion",
                    "significado",
                    "id",
                ):
                    if keyword in terms.split():
                        score += 5

                if score > best_score:
                    best_score = score
                    best_row = row_number
                    best_values = headers

            result[sheet_name] = {
                "header_row": best_row,
                "headers": best_values,
            }

        return result
    finally:
        workbook.close()


def sugerir_mapeo(
    headers: list[str],
    schema_path: str | Path,
    threshold: float = 0.55,
) -> list[SugerenciaEncabezado]:
    schema = cargar_esquema(schema_path)
    suggestions: list[SugerenciaEncabezado] = []
    candidates: list[tuple[str, str]] = []

    for field in schema.campos:
        candidates.append(
            (field.nombre_canonico, field.nombre_canonico)
        )

        for alias in field.aliases:
            candidates.append(
                (field.nombre_canonico, alias)
            )

    for header in headers:
        best_field = None
        best_alias = None
        best_score = 0.0

        for field_name, alias in candidates:
            score = similitud(header, alias)

            if score > best_score:
                best_field = field_name
                best_alias = alias
                best_score = score

        if best_score >= 0.90:
            decision = "automatico_alta_confianza"
        elif best_score >= threshold:
            decision = "propuesto_para_revision"
        else:
            decision = "sin_correspondencia"

        suggestions.append(
            SugerenciaEncabezado(
                original=header,
                normalizado=normalizar_texto(header),
                campo_canonico=(
                    best_field
                    if best_score >= threshold
                    else None
                ),
                alias_recomendado=(
                    best_alias
                    if best_score >= threshold
                    else None
                ),
                confianza=best_score,
                decision=decision,
            )
        )

    return suggestions


def construir_aliases_aprobados(
    suggestions: list[SugerenciaEncabezado],
    minimum_confidence: float = 0.90,
) -> dict[str, str]:
    return {
        item.original: item.alias_recomendado
        for item in suggestions
        if item.confianza >= minimum_confidence
        and item.alias_recomendado
    }


def extender_esquema_temporal(
    schema_path: str | Path,
    output_path: str | Path,
    approved_aliases: dict[str, str],
) -> Path:
    source = Path(schema_path)
    target = Path(output_path)

    data = json.loads(
        source.read_text(encoding="utf-8")
    )

    fields = data.get("fields", [])
    canonical_by_alias: dict[str, str] = {}

    for field in fields:
        canonical = str(field["name"])
        canonical_by_alias[normalizar_texto(canonical)] = canonical

        for alias in field.get("aliases", []):
            canonical_by_alias[
                normalizar_texto(str(alias))
            ] = canonical

    for original, recommended_alias in approved_aliases.items():
        canonical = canonical_by_alias.get(
            normalizar_texto(recommended_alias)
        )

        if canonical is None:
            continue

        for field in fields:
            if field["name"] != canonical:
                continue

            aliases = field.setdefault("aliases", [])

            if original not in aliases:
                aliases.append(original)

            break

    data["version"] = (
        str(data.get("version", "1.0.0"))
        + "-p07-normalized"
    )
    data["generated_by"] = "SPT-001B-P07-v1.1"

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(
            data,
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    return target


def ejecutar_normalizacion(
    *,
    excel: str | Path,
    schema: str | Path,
    output_dir: str | Path,
) -> dict[str, Any]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)

    detected = leer_encabezados_excel(excel)

    all_headers: list[str] = []

    for sheet_data in detected.values():
        all_headers.extend(sheet_data["headers"])

    suggestions = sugerir_mapeo(
        all_headers,
        schema,
    )

    approved = construir_aliases_aprobados(suggestions)

    mapping_path = output / "header-mapping-analysis.json"
    mapping_path.write_text(
        json.dumps(
            {
                "increment": "SPT-001B-P07",
                "version": "1.1.0",
                "detected": detected,
                "suggestions": [
                    asdict(item)
                    for item in suggestions
                ],
                "approved_aliases": approved,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    normalized_schema = extender_esquema_temporal(
        schema,
        output / "schema-p07-normalized.json",
        approved,
    )

    contract = cargar_esquema(normalized_schema)
    reading = LectorExcelRLB(contract).leer(excel)

    artifacts = exportar_resultado(
        reading,
        output / "reprocessed",
    )

    if reading.perfil is None:
        raise RuntimeError(
            "No se generó perfil después de normalizar."
        )

    result = {
        "mapping_analysis": mapping_path,
        "normalized_schema": normalized_schema,
        "artifacts": artifacts,
        "profile": reading.perfil,
    }

    summary_path = output / "normalization-summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "increment": "SPT-001B-P07",
                "version": "1.1.0",
                "total_headers": len(all_headers),
                "approved_aliases": len(approved),
                "total_records": reading.perfil.total_registros,
                "valid_records": (
                    reading.perfil.total_registros_validos
                ),
                "invalid_records": (
                    reading.perfil.total_registros_con_errores
                ),
                "mapping_analysis": str(mapping_path),
                "normalized_schema": str(normalized_schema),
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    result["summary"] = summary_path
    return result


def construir_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Analiza y normaliza encabezados del RLB oficial."
        )
    )
    parser.add_argument("--excel", required=True)
    parser.add_argument(
        "--schema",
        default="config/rlb/schema-v1.json",
    )
    parser.add_argument(
        "--output",
        default="artifacts/rlb/SPT-001B-P07",
    )
    return parser


def main() -> int:
    args = construir_parser().parse_args()

    result = ejecutar_normalizacion(
        excel=args.excel,
        schema=args.schema,
        output_dir=args.output,
    )

    profile = result["profile"]

    print("SPT-001B-P07 ejecutado correctamente.")
    print(f"Registros: {profile.total_registros}")
    print(f"Válidos: {profile.total_registros_validos}")
    print(
        "Con errores: "
        f"{profile.total_registros_con_errores}"
    )
    print(f"Resumen: {result['summary']}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())