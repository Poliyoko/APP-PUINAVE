"""Lectura segura y trazable del Repositorio Léxico Base en Excel."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .models import RegistroLexico
from .profile_models import PerfilHojaRLB, PerfilRepositorioRLB
from .schema import EsquemaRLB


@dataclass(slots=True)
class ErrorFilaRLB:
    """Error asociado con una fila concreta del Excel."""

    hoja: str
    fila: int
    mensajes: list[str]


@dataclass(slots=True)
class ResultadoLecturaRLB:
    """Resultado completo de una importación del RLB."""

    registros: list[RegistroLexico] = field(default_factory=list)
    errores: list[ErrorFilaRLB] = field(default_factory=list)
    perfil: PerfilRepositorioRLB | None = None


class LectorExcelRLB:
    """Lector institucional del Repositorio Léxico Base."""

    def __init__(
        self,
        esquema: EsquemaRLB,
        *,
        max_filas_busqueda_encabezado: int = 20,
    ) -> None:
        if max_filas_busqueda_encabezado < 1:
            raise ValueError(
                "max_filas_busqueda_encabezado debe ser mayor que cero."
            )

        self.esquema = esquema
        self.max_filas_busqueda_encabezado = (
            max_filas_busqueda_encabezado
        )

    def leer(
        self,
        ruta_excel: str | Path,
        *,
        hojas: Iterable[str] | None = None,
    ) -> ResultadoLecturaRLB:
        """Lee el Excel y conserva trazabilidad por archivo, hoja y fila."""

        ruta = Path(ruta_excel)

        if not ruta.is_file():
            raise FileNotFoundError(
                f"No se encontró el archivo RLB: {ruta}"
            )

        if ruta.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError(
                "El RLB debe utilizar formato .xlsx o .xlsm."
            )

        workbook = load_workbook(
            filename=ruta,
            read_only=True,
            data_only=True,
        )

        try:
            nombres_hojas = list(hojas or workbook.sheetnames)

            faltantes = [
                nombre
                for nombre in nombres_hojas
                if nombre not in workbook.sheetnames
            ]

            if faltantes:
                raise ValueError(
                    "No existen las siguientes hojas: "
                    + ", ".join(faltantes)
                )

            perfil = PerfilRepositorioRLB(
                archivo=ruta.name,
                version_esquema=self.esquema.version,
                total_hojas=len(nombres_hojas),
            )

            resultado = ResultadoLecturaRLB(perfil=perfil)

            for nombre_hoja in nombres_hojas:
                hoja = workbook[nombre_hoja]
                perfil_hoja = self._procesar_hoja(
                    hoja=hoja,
                    archivo=ruta.name,
                    resultado=resultado,
                )
                perfil.hojas.append(perfil_hoja)

            perfil.total_registros = sum(
                hoja.total_registros
                for hoja in perfil.hojas
            )
            perfil.total_registros_validos = sum(
                hoja.total_registros_validos
                for hoja in perfil.hojas
            )
            perfil.total_registros_con_errores = sum(
                hoja.total_registros_con_errores
                for hoja in perfil.hojas
            )

            return resultado
        finally:
            workbook.close()

    def _procesar_hoja(
        self,
        *,
        hoja: Any,
        archivo: str,
        resultado: ResultadoLecturaRLB,
    ) -> PerfilHojaRLB:
        fila_encabezado, encabezados = self._detectar_encabezados(hoja)

        perfil = PerfilHojaRLB(
            nombre=hoja.title,
            fila_encabezado=fila_encabezado,
            total_filas_fisicas=hoja.max_row,
            total_columnas_fisicas=hoja.max_column,
        )

        if fila_encabezado is None:
            perfil.errores.append(
                "No fue posible identificar una fila de encabezados."
            )
            return perfil

        perfil.columnas = encabezados

        valores_por_columna = {
            encabezado: 0
            for encabezado in encabezados
        }

        conocidas: set[str] = set()
        desconocidas: set[str] = set()

        for numero_fila, valores in enumerate(
            hoja.iter_rows(
                min_row=fila_encabezado + 1,
                values_only=True,
            ),
            start=fila_encabezado + 1,
        ):
            if self._fila_vacia(valores):
                continue

            fila = {
                encabezado: (
                    valores[indice]
                    if indice < len(valores)
                    else None
                )
                for indice, encabezado in enumerate(encabezados)
            }

            for columna, valor in fila.items():
                if valor not in (None, ""):
                    valores_por_columna[columna] += 1

            mapeo = self.esquema.mapear_fila(
                fila,
                archivo=archivo,
                hoja=hoja.title,
                numero_fila=numero_fila,
            )

            perfil.total_registros += 1
            conocidas.update(mapeo.columnas_reconocidas.keys())
            desconocidas.update(mapeo.columnas_desconocidas)

            if mapeo.errores:
                perfil.total_registros_con_errores += 1
                resultado.errores.append(
                    ErrorFilaRLB(
                        hoja=hoja.title,
                        fila=numero_fila,
                        mensajes=mapeo.errores,
                    )
                )
            else:
                perfil.total_registros_validos += 1

            resultado.registros.append(mapeo.registro)

        perfil.columnas_reconocidas = sorted(conocidas)
        perfil.columnas_desconocidas = sorted(desconocidas)
        perfil.columnas_vacias = sorted(
            columna
            for columna, total in valores_por_columna.items()
            if total == 0
        )

        return perfil

    def _detectar_encabezados(
        self,
        hoja: Any,
    ) -> tuple[int | None, list[str]]:
        mejor_fila = None
        mejores_encabezados: list[str] = []
        mejor_puntaje = 0

        limite = min(
            hoja.max_row,
            self.max_filas_busqueda_encabezado,
        )

        for numero_fila, valores in enumerate(
            hoja.iter_rows(
                min_row=1,
                max_row=limite,
                values_only=True,
            ),
            start=1,
        ):
            encabezados = self._normalizar_encabezados(valores)
            no_vacios = [
                valor
                for valor in encabezados
                if valor.strip()
            ]

            if not no_vacios:
                continue

            puntaje = len(no_vacios)
            terminos = " ".join(no_vacios).lower()

            for termino in (
                "puinave",
                "español",
                "espanol",
                "ingles",
                "inglés",
                "palabra",
                "traduccion",
                "traducción",
            ):
                if termino in terminos:
                    puntaje += 5

            if puntaje > mejor_puntaje:
                mejor_puntaje = puntaje
                mejor_fila = numero_fila
                mejores_encabezados = encabezados

        if mejor_fila is None:
            return None, []

        return (
            mejor_fila,
            self._hacer_encabezados_unicos(mejores_encabezados),
        )

    @staticmethod
    def _normalizar_encabezados(
        valores: tuple[Any, ...],
    ) -> list[str]:
        encabezados = []

        for indice, valor in enumerate(valores, start=1):
            if valor is None or not str(valor).strip():
                encabezados.append(
                    f"columna_sin_nombre_{indice}"
                )
            else:
                encabezados.append(str(valor).strip())

        return encabezados

    @staticmethod
    def _hacer_encabezados_unicos(
        encabezados: list[str],
    ) -> list[str]:
        usados: dict[str, int] = {}
        resultado: list[str] = []

        for encabezado in encabezados:
            total = usados.get(encabezado, 0) + 1
            usados[encabezado] = total

            if total == 1:
                resultado.append(encabezado)
            else:
                resultado.append(
                    f"{encabezado}__{total}"
                )

        return resultado

    @staticmethod
    def _fila_vacia(
        valores: tuple[Any, ...],
    ) -> bool:
        return all(
            valor is None
            or (
                isinstance(valor, str)
                and not valor.strip()
            )
            for valor in valores
        )